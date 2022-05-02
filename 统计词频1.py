# 加载txt列表寻找关键词并保存到excel
from openpyxl import Workbook
import openpyxl
import re
import os
import shutil
import datetime
starttime = datetime.datetime.now()

def initialization():
    wb = Workbook()
    # 在索引为0的位置创建一个名为mytest的sheet页
    ws = wb.create_sheet('年报关键词词频统计',0)
    # 对sheet页设置一个颜色（16位的RGB颜色）
    ws.sheet_properties.tabColor = 'ff72BA'
    row=['股票代码','年份',keywords[0],keywords[1],keywords[2],keywords[3],keywords[4],keywords[5],"总词数"]
    ws.append(row)
    wb.save('D:/年报关键词词频统计.xlsx')
    wb.close()



def matchKeyWords(txt_folder, keyWords, code):
    files = os.listdir(txt_folder)
    words_num = []  # 保存所有文件词频
    L=[]
    for file in files:
        word_freq = {}  # 单词出现频率次：word：num
        if os.path.splitext(file)[-1] == ".txt":
            txt_path = os.path.join(txt_folder, file)
            with open(txt_path, "r", encoding='utf-8', errors='ignore')as fp:
                text = fp.readlines()
                for word in keyWords:
                    num = 0
                    for line in text:
                        num += line.count(word)
                    word_freq[word] = num
                stock_code = code
                year = re.sub("\D", "", file)
                stock_name = ""
                words_num.append((word_freq, stock_code, stock_name,year))
                lens = 0
                for line in text:
                    l = re.sub('[^\u4e00-\u9fa5]+','',line)
                    lens += len(l)
                L.append(lens)

    wb = openpyxl.load_workbook('D:/年报关键词词频统计.xlsx')
    ws = wb['年报关键词词频统计']
    i = 0
    for index, one in enumerate(words_num):
        row=[]
        word_f = one[0]
        stock_code = one[1]
        stock_name = one[2]
        year = one[3]
        row.append(stock_code)
        row.append(year)
        row.append(word_f[keywords[0]])
        row.append(word_f[keywords[1]])
        row.append(word_f[keywords[2]])
        row.append(word_f[keywords[3]])
        row.append(word_f[keywords[4]])
        row.append(word_f[keywords[5]])
        row.append(str(L[i]))
        ws.append(row)
        i = i +1

    wb.save('D:/年报关键词词频统计.xlsx')
    wb.close()

def mk(origindir,keywords1):
    dir_list_1 = os.listdir(origindir)
    for dir in dir_list_1:
        dir_1 = origindir + dir + '/'
        if os.path.isdir(dir_1):
            cod = re.sub("\D", "", dir)
            matchKeyWords(dir_1,keywords1,cod)
            print("正在计算{}".format(cod))



folder = 'D:/final/'
keywords = ['大数据','海量数据','数据中心','信息资产','数据化','算力']
initialization()

mk(folder, keywords)
